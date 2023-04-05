import openpyxl
from empresas import empresas
import yfinance as yf
from datetime import datetime
import calendar
import pandas as pd
from playsound import playsound
import os
import glob


def obter_ultimo_dia_do_mes(ano, mes):
    """Obtém o último dia do mês"""
    return calendar.monthrange(ano, mes)[1]


def obter_data_fim_disponivel_para_acao(acao, ano, mes):
    """Obtém a última data disponível para a ação no mês desejado"""
    data = yf.download(
        acao, start=f"{ano}-{mes}-01", end=f"{ano}-{mes}-{obter_ultimo_dia_do_mes(ano, mes)}")
    if data.empty:
        return None
    data = data.dropna()
    return data.index[-1].strftime('%Y-%m-%d')


def obter_dados_do_periodo(acao, ano, mes):
    """Obtém os dados do período desejado"""
    data_inicio = f"{ano}-{mes}-01"
    data_fim = obter_data_fim_disponivel_para_acao(acao, ano, mes)
    dados = yf.download(acao, start=data_inicio, end=data_fim)
    return dados


def calcular_retorno_e_variacao_percentual(dados):
    """Calcula o retorno e a variação percentual"""
    preco_inicio = dados['Close'][0]
    preco_fim = dados['Close'][-1]
    retorno = (preco_fim - preco_inicio) / preco_inicio
    variacao = round(retorno * 100, 2)
    return round(retorno, 2), variacao


def gerar_relatorio(acao, ano, mes):
    """Gera um relatório para a ação e mês desejados"""
    dados = obter_dados_do_periodo(acao, ano, mes)
    if dados.empty:
        return None
    retorno, variacao = calcular_retorno_e_variacao_percentual(dados)
    data_inicio = f"{ano}-{mes}-01"
    data_fim = obter_data_fim_disponivel_para_acao(acao, ano, mes)
    return {
        'ano': ano,
        'acao': acao,
        'retorno': retorno,
        'variacao_percentual': variacao,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
    }


def gerar_relatorio_para_meses_anteriores(acao, n_anos):
    """Gera um relatório para a ação nos últimos n_anos"""
    relatorios = []
    mes_atual = datetime.now().month
    ano_atual = datetime.now().year
    first_year = True

    for i in range(ano_atual - n_anos, ano_atual):
        relatorio = gerar_relatorio(acao, i, mes_atual)

        if first_year:
            if relatorio is None:
                break
            first_year = False

        if relatorio is not None and 'retorno' in relatorio:
            relatorios.append(relatorio)

    df = pd.DataFrame(relatorios)
    if not df.empty:
        df = df.pivot_table(index='acao', columns='ano', values='retorno')
    return df


def salvar_relatorios_em_excel():
    """Salva os relatórios em um arquivo Excel"""
    n_anos = 10
    dfs = []
    empresas_sem_relatorio = []

    # Cria uma nova guia para empresas sem relatório
    wb = openpyxl.Workbook()
    ws_empresas_sem_relatorio = wb.active
    ws_empresas_sem_relatorio.title = "Empresas sem relatório"
    ws_empresas_sem_relatorio.cell(
        row=1, column=1, value="Empresas sem relatório")

    for empresa in empresas:
        relatorio = gerar_relatorio_para_meses_anteriores(empresa, n_anos)
        if not relatorio.empty:
            dfs.append(relatorio)
        else:
            empresas_sem_relatorio.append(empresa)
            ws_empresas_sem_relatorio.append([empresa])

    if dfs:
        result = pd.concat(dfs)
        data_atual = datetime.now().strftime("%Y_%m_%d")
        nome_arquivo = f'relatorios_empresas_{data_atual}.xlsx'

        # Excluir o arquivo anterior
        arquivos_antigos = glob.glob('relatorios_empresas_*.xlsx')
        for arquivo in arquivos_antigos:
            os.remove(arquivo)

        result.to_excel(nome_arquivo)

        # Abrir o arquivo Excel e adicionar a data e hora de geração da planilha
        wb = openpyxl.load_workbook(nome_arquivo)
        ws = wb.active
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=1, column=ws.max_column + 2,
                value=f"Planilha gerada em: {current_time}")

        # Mover a célula com a data e hora de geração para o topo da planilha
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value=ws.cell(
            row=2, column=ws.max_column).value)
        ws.cell(row=2, column=ws.max_column).value = None

        # Congelar as duas primeiras linhas
        ws.freeze_panes = "A3"

        # Exclui as empresas que não geraram relatório do arquivo empresas.py
        for empresa in empresas_sem_relatorio:
            empresas.remove(empresa)

        # Salva a lista atualizada de empresas no arquivo empresas.py
        with open('empresas.py', 'w') as f:
            f.write(f"empresas = {empresas}")

    else:
        # Salvar a planilha mesmo se não houver relatórios
        data_atual = datetime.now().strftime("%Y_%m_%d")
        nome_arquivo = f'relatorios_empresas_{data_atual}.xlsx'

        # Excluir o arquivo anterior
    arquivos_antigos = glob.glob('relatorios_empresas_*.xlsx')
    for arquivo in arquivos_antigos:
        os.remove(arquivo)

    result.to_excel(nome_arquivo)

    # Adicionar uma nova guia (planilha) para as empresas sem relatório
    if empresas_sem_relatorio:
        ws_empresas_sem_relatorio = wb.create_sheet(
            "Empresas sem relatório")
        for index, empresa in enumerate(empresas_sem_relatorio, start=1):
            ws_empresas_sem_relatorio.cell(
                row=index, column=1, value=empresa)

    wb.save(nome_arquivo)
    wb.close()

    print(f"Arquivo '{nome_arquivo}' criado com sucesso.")

    # Remover as empresas que não geraram relatório da lista de empresas
    for empresa in empresas_sem_relatorio:
        if empresa in empresas:
            empresas.remove(empresa)
        else:
            print(f"{empresa} não encontrado na lista de empresas.")

    # Salva a lista atualizada de empresas no arquivo empresas.py
    with open('empresas.py', 'w') as f:
        f.write(f"empresas = {empresas}")


# Chamar a função para salvar os relatórios das empresas em um arquivo Excel
salvar_relatorios_em_excel()

# Reproduzir um som para indicar que a pesquisa foi concluída
# playsound("audios/success.wav")
